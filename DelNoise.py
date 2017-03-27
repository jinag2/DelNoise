#! python3

import csv
import sys
import os
import copy
from openpyxl import Workbook
from openpyxl import styles
from openpyxl import chart
from openpyxl import drawing
from openpyxl.drawing import line


MAX_ADC_NUM = 10
CUT_RANGE = int(MAX_ADC_NUM / 3)
LINE_COLOR = ('416FA6', 'A8423F', '86A44A', '6E548D', '3D96AE',
              'B8860B', 'E9967A', 'DA8137', '8EA5CB', '808000',
              'A0522D', '2E8B57', 'B0E0E6', '000080', 'FFDEAD')
MEDIAN_COLOR = 'FF0000'
FILTER_COLOR = '0000FF'


def main():
    csv_list = read_csv(sys.argv[1])
    voltage_list = abstract_float_col(csv_list, 0)
    
    if len(csv_list[0]) - 4 > MAX_ADC_NUM:  # 4: voltage column and tree flag columns
        adc_num = MAX_ADC_NUM
    else:
        adc_num = len(csv_list[0]) - 4
    
    adc_lists = []
    if adc_num >= 1:
        for i in range(adc_num):
            adc_list = abstract_float_col(csv_list, i+1)
            adc_lists.append(adc_list)
            # data_list.append(adc_list)
    else:
        print('No ADC data.')
        sys.exit()
        
    # csv_out = cut_flag(adc_list)
    # output_file = write_csv(csv_out)
    # print('Creat ' + output_file + ' OK!')

    median_list = calc_median(adc_lists)
    average_lists = calc_average(adc_lists, CUT_RANGE)
    excel_file = generate_excel_file(sys.argv[1])
    if create_excel(excel_file, voltage_list, adc_lists, median_list, average_lists):
        print('Create ' + excel_file + ' succeeded!')
    else:
        print('Create ' + excel_file + ' failed!')


def read_csv(filename):
    try:
        csv_file = open(filename)
    except Exception as err:
        print('Can not open ' + ' (' + str(err) + ')')
        sys.exit()
        
    csv_reader = csv.reader(csv_file)
    return list(csv_reader)


def abstract_float_col(data_list, col):
    value_list = []
    for row in data_list:
        value_list.append(float(row[col]))
    return value_list


# def abstract_int_col(data_list, col):
#     value_list = []
#     for row in data_list:
#         value_list.append(int(row[col]))
#     return value_list


def cut_flag(data_list):
    flag_size = len(data_list[0])
    new_list = copy.deepcopy(data_list)
    for row in new_list:
        del row[flag_size - 1]
        del row[flag_size - 2]
        del row[flag_size - 3]
    return new_list


def write_csv(data_list):
    input_file_tuple = os.path.split(sys.argv[1])
    file_tuple = os.path.splitext(input_file_tuple[1])
    filename = input_file_tuple[0] + '\\' + file_tuple[0] + '_noflag' + file_tuple[1]
    
    csv_file = open(filename, 'w', newline = '')
    csv_writer = csv.writer(csv_file)
    for row in data_list:
        csv_writer.writerow(row)
    csv_file.close()
    
    return filename


def calc_median(adc_list):
    adc_num = len(adc_list)
    odd_num = True
    odd_index = int(adc_num // 2)
    if (adc_num % 2) == 0:
        odd_num = False
    
    median_list = []
    for i in range(len(adc_list[0])):
        elem_list = []
        for j in range(adc_num):
            elem_list.append(adc_list[j][i])
        
        elem_list.sort()
        if odd_num:
            median_list.append(elem_list[odd_index])
        else:
            elem = round((elem_list[odd_index - 1] + elem_list[odd_index]) / 2)
            median_list.append(elem)
    
    return median_list


def calc_average(adc_list, cut_range):
    adc_num = len(adc_list)
    if cut_range > (adc_num / 3):
        max_cut = int(adc_num / 3)
    else:
        max_cut = int(cut_range)
        
    total_lists = []
    for cut in range(max_cut + 1):
        average_list = []
        for i in range(len(adc_list[0])):
            elem_sum = 0
            for j in range(cut, adc_num - cut):
                elem_sum += adc_list[j][i]
                
            average_list.append(elem_sum / (adc_num - cut * 2))
        total_lists.append(average_list)

    return total_lists


def generate_excel_file(csv_filename):
    input_file_tuple = os.path.split(csv_filename)
    file_tuple = os.path.splitext(input_file_tuple[1])
    filename = input_file_tuple[0] + '\\' + file_tuple[0] + '_out.xlsx'
    return filename


def create_scatter_chart(sheet, title, y_axis_title, x_axis_title,
                         start_item, item_size, serials_num, line_color_list):
    adc_chart = chart.ScatterChart()
    adc_chart.title = title
    adc_chart.style = 13
    adc_chart.y_axis.title = y_axis_title
    adc_chart.x_axis.title = x_axis_title
    adc_chart.height = 12
    adc_chart.width = 16

    min_row_idx = start_item + 1
    max_row_idx = min_row_idx + item_size - 1
    x_values = chart.Reference(sheet, min_col = 1, min_row = min_row_idx, max_row = max_row_idx)
    for i in range(serials_num):
        column_idx = i + 2
        values = chart.Reference(sheet, min_col = column_idx, min_row = min_row_idx, max_row = max_row_idx)
        series = chart.Series(values, x_values, title = sheet.cell(row = 1, column = column_idx).value)
        if i == serials_num - 1:
            series.graphicalProperties.line = drawing.line.LineProperties(solidFill = line_color_list[serials_num-1])
        else:
            series.graphicalProperties.line = drawing.line.LineProperties(solidFill = line_color_list[i])
        series.graphicalProperties.line.width = 27432  # width in EMUs, EMU = pixel * 914400 / 96, assume pixel = 75
        adc_chart.series.append(series)
        
    return adc_chart

    
def create_excel(filename, voltage_list, adc_lists, median_list, average_lists):
    wb = Workbook()
    data_sheet = wb.active
    data_sheet.title = "Data"
    font12 = styles.Font(size = 12)
    data_sheet.column_dimensions['A'].width = 11

    serial_title = ['Voltage']
    adc_num = len(adc_lists)
    for i in range(adc_num):
        serial_title.append('ADC ' + str(i))
 
    serial_title.append('Median')
    
    average_num = len(average_lists)
    for i in range(average_num):
        serial_title.append('Average ' + str(i))

    data_sheet.append(serial_title)
    
    for row_idx in range(len(voltage_list)):
        cell = data_sheet.cell(row = row_idx + 2, column = 1, value = voltage_list[row_idx])
        cell.font = font12
        
    for col_idx in range(len(adc_lists)):
        for row_idx in range(len(adc_lists[col_idx])):
            cell = data_sheet.cell(row = row_idx + 2, column = col_idx + 2, value = adc_lists[col_idx][row_idx])
            cell.font = font12

    for row_idx in range(len(median_list)):
        cell = data_sheet.cell(row = row_idx + 2, column = adc_num + 2, value = median_list[row_idx])
        cell.font = font12

    for col_idx in range(len(average_lists)):
        for row_idx in range(len(average_lists[col_idx])):
            cell = data_sheet.cell(row = row_idx + 2, column = col_idx + adc_num + 3, value = round(average_lists[col_idx][row_idx]))
            cell.font = font12
            
    chart_sheet = wb.create_sheet(index = 1, title = "Chart")

    # ADC chart
    line_color_list = []
    for i in range(adc_num):
        line_color_list.append(LINE_COLOR[i])

    line_color_list.append(MEDIAN_COLOR)
    
    serials_num = adc_num + 1
    adc_chart = create_scatter_chart(data_sheet, "ADC Chart", 'ADC', 'Voltage',
                                     1, 200, serials_num, line_color_list)
    chart_sheet.add_chart(adc_chart, "B1")

    # Average chart
    line_color_list = []
    for i in range(average_num):
        line_color_list.append(LINE_COLOR[i])

    line_color_list.append(MEDIAN_COLOR)

    serials_num = average_num + 1
    adc_chart = create_scatter_chart(data_sheet, "Average Chart", 'ADC', 'Voltage',
                                     1, 200, serials_num, line_color_list)
    chart_sheet.add_chart(adc_chart, "L1")

    try:
        wb.save(filename)
    except Exception as err:
        print('Can not save ' + ' (' + str(err) + ')')
        sys.exit()
        
    return True


main()
